import os

import openpyxl as xl
from openpyxl.chart import LineChart, Reference, BarChart,  PieChart
from openpyxl import Workbook

from openpyxl.chart.marker import DataPoint

from openpyxl.styles import Font, Alignment, Side, Border, PatternFill
import openpyxl.styles.numbers

from docxtpl import DocxTemplate

# whatis = lambda obj: print(type(obj), "\n\t" + "\n\t".join(dir(obj)))

from openpyxl.styles.borders import BORDER_THIN

def create_a_report(path_load, path_save, week):

    workbook = xl.load_workbook(path_load,
                                data_only=True)  # открываем базу с отчетами, data_only=True - только данные (без формул)
    sheet_1 = workbook.active  # выбираем активный лист или sheet_1 = workbook['Отчет для АО НИПОМ']  # выбираем нужный лист
    wb = Workbook()  # создаем рабочую книгу, в которую будем сохранять данные из workbook
    ws = wb.active  # создаем рабочий лист
    ws.sheet_properties.tabColor = "1072BA"  # задаем цвет вкладки
    ws.title = "Графики"  # задаем имя вкладки

    # задаем наименование столбцов
    ws['A1'] = "Неделя"
    ws['B1'] = "Wгпэг,\nкВт*ч"
    ws['C1'] = "Wсм,\nкВт*ч"
    ws['D1'] = "Wсумм,\nкВт*ч"
    ws['E1'] = "Wпотр,\nкВт*ч"
    ws['F1'] = "Wсн,\nкВт*ч"
    ws['G1'] = "Моточасы\n(ГПЭГ)"
    ws['H1'] = "Vст (Газ),\nм^3"
    ws['I1'] = "Qкотла,\nкВт*ч"
    ws['J1'] = "Vгпэг (Газ),\nм^3"
    ws['K1'] = "Vкотел (Газ),\nм^3"
    ws['L1'] = "Vгпэг/Wсумм,\nм3/кВт*ч"
    ws['M1'] = "T_min,\n°C"
    ws['N1'] = "T_max,\n°C"
    ws['O1'] = "Дата начала"
    ws['P1'] = "Дата\nокончания"

    i = 2  # переменная для итерации строк в последующем цикле for
    # for row in range(10073, sheet_1.max_row + 1):  # цикл по строкам, начиная с нужной

    week_OPE = [50, 51, 52, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26,
                27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45, 46, 47, 48, 49,
                50, 51, 52, 53]  # количество недель ОПЭ
    count = 0
    for row in range(5034, sheet_1.max_row + 1):  # цикл по строкам с данными, начиная с 50-й недели
        if sheet_1.cell(row, 46).value == week_OPE[i - 2]:  # если значение в ячейке
            # равно номеру недели из списка
            t_min = ws.cell(i, 13)  # создаем переменную минимальной температуры АБ
            t_max = ws.cell(i, 14)  # создаем переменную максимальной температуры АБ
            data_start = ws.cell(i, 15)
            if t_min.value is None:  # проверка на наличии в ячейке значения
                t_min.value = 100
            if t_max.value is None:  # проверка на наличии в ячейке значения
                t_max.value = 0
            if sheet_1.cell(row, 23).value < t_min.value:  # проверка на меньшее значение
                t_min.value = sheet_1.cell(row, 23).value
            if sheet_1.cell(row, 23).value > t_max.value:  # проверка на большее значение
                t_max.value = sheet_1.cell(row, 23).value
            if count == 0:
                data_start.value = sheet_1.cell(row, 3).value
                count = 1
        if sheet_1.cell(row, 46).value == week_OPE[i - 2] and sheet_1.cell(row, 47).value == 0:  # если значение в ячейке
            # равно номеру недели из списка и значение расхода не равно нулю
            v_kotel = ws.cell(i, 11)  # создаем переменную расхода котла без учета времени работы ГПЭГ
            if v_kotel.value is None:  # проверка на наличии в ячейке значения
                v_kotel.value = 0
            v_kotel.value += sheet_1.cell(row, 48).value  # суммирование значений в ячейках
            if ws.cell(i, 20).value is None:  # проверка на наличии в ячейке значения
                ws.cell(i, 20).value = 0
            ws.cell(i, 20).value += 1  # суммирование количества ячеек

        if sheet_1.cell(row, 1).value is not None:  # проверяем пустая ячейка первого столбца или нет
            # print(sheet_1.cell(row, 1).value)
            week_cell = ws.cell(i, 1)  # создаем переменную строк для столбца с номерами недель
            power_GPEG = ws.cell(i, 2)  # создаем переменную строк для столбца с выработанной мощности ГПЭГ
            power_Sun = ws.cell(i, 3)  # создаем переменную строк для столбца с выработанной мощности солнечного модуля
            power_Sum = ws.cell(i, 4)  # создаем переменную строк для столбца с выработанной суммарной мощностью
            power_Potr = ws.cell(i, 5)  # создаем переменную строк для столбца с потребленной нагрузкой мощностью
            power_SN = ws.cell(i, 6)  # создаем переменную строк для столбца с собственными нуждами
            mototime_GPEG = ws.cell(i, 7)  # создаем переменную строк для столбца с моточасами ГПЭГ
            v_Sum = ws.cell(i, 8)  # создаем переменную строк для столбца с общим потребленным объемом газа
            data_end = ws.cell(i, 16)

            week_cell.value = sheet_1.cell(row, 1).value  # присваиваем переменной значение из базовой таблицы
            week_cell.number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[1]  # формат ячейки 0
            power_GPEG.value = sheet_1.cell(row, 29).value  # присваиваем переменной значение из базовой таблицы
            power_GPEG.number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[2]  # формат ячейки 0.00
            power_Sun.value = sheet_1.cell(row, 30).value  # присваиваем переменной значение из базовой таблицы
            power_Sun.number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[2]  # формат ячейки 0.00
            power_Sum.value = sheet_1.cell(row, 31).value  # присваиваем переменной значение из базовой таблицы
            power_Sum.number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[2]  # формат ячейки 0.00
            power_Potr.value = sheet_1.cell(row, 32).value  # присваиваем переменной значение из базовой таблицы
            power_Potr.number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[2]  # формат ячейки 0.00
            power_SN.value = sheet_1.cell(row, 33).value  # присваиваем переменной значение из базовой таблицы
            power_SN.number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[2]  # формат ячейки 0.00
            mototime_GPEG.value = sheet_1.cell(row, 34).value  # присваиваем переменной значение из базовой таблицы
            mototime_GPEG.number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[2]  # формат ячейки 0.00
            v_Sum.value = sheet_1.cell(row, 35).value  # присваиваем переменной значение из базовой таблицы
            v_Sum.number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[2]  # формат ячейки 0.00
            data_end.value = sheet_1.cell(row, 3).value
            count = 0

            i += 1
    for i in range(2, ws.max_row + 1):
        if ws.cell(i, 1).value is not None:
            ws.cell(i, 11).value = ws.cell(i, 11).value / (
                    ws.cell(i, 20).value / 30) * 168  # столбец с данными расхода газа котлом
            ws.cell(i, 11).number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[2]  # формат ячейки 0.00
            ws.cell(i, 10).value = ws.cell(i, 8).value - ws.cell(i, 11).value  # столбец с данными расхода газа ГПЭГ
            ws.cell(i, 10).number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[2]  # формат ячейки 0.00
            ws.cell(i, 9).value = ws.cell(i, 11).value * 9.5  # столбец с данными выработки тепловой энергии котла
            ws.cell(i, 9).number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[1]  # формат ячейки 0
            ws.cell(i, 12).value = ws.cell(i, 10).value / ws.cell(i,
                                                                  4).value  # расчет эффективности выработки БКЭУ, выраженной через Vгпэг / Wсумм
            ws.cell(i, 12).number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[2]  # формат ячейки 0.00
            ws.cell(i, 20).value = None  # удаляем временное значение количества ячеек
        else:
            ws.cell(i, 11).value = None  # удаляем значения выходящие за пределы расчетных недель
            ws.cell(i, 13).value = None  # удаляем значения выходящие за пределы расчетных недель
            ws.cell(i, 14).value = None  # удаляем значения выходящие за пределы расчетных недель
            ws.cell(i, 15).value = None  # удаляем значения выходящие за пределы расчетных недель
            ws.cell(i, 20).value = None  # удаляем значения выходящие за пределы расчетных недель

    # задаем ширину столбцов
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 14
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 10
    ws.column_dimensions['N'].width = 10
    ws.column_dimensions['O'].width = 18
    ws.column_dimensions['P'].width = 18

    thin_border = Border(  # выделение границ ячеек
        left=Side(border_style=BORDER_THIN, color='00000000'),
        right=Side(border_style=BORDER_THIN, color='00000000'),
        top=Side(border_style=BORDER_THIN, color='00000000'),
        bottom=Side(border_style=BORDER_THIN, color='00000000')
    )

    # цикл для задания ячейкам заголовков свойств
    for row in ws.iter_cols(min_col=1, max_col=16, min_row=1, max_row=1):
        for cel in row:
            cel.font = Font(size=12, bold=True)  # размер шрифта и жирное выделение
            cel.alignment = Alignment(horizontal="center", vertical="center",
                                      wrapText=True)  # выравнивание по центру и разрешение переноса строк
            cel.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")

    # цикл для выделения границ ячеек
    for row in ws.iter_cols(min_col=1, max_col=16, min_row=1, max_row=ws.max_row):
        for cel in row:
            cel.border = thin_border
    # print(week)
    # print(ws.cell(ws.max_row-1, 1).value)
    if ws.cell(ws.max_row-1, 1).value < week:
        return "Задана неделя вне диапазона текущего ОПЭ, равного {} недель в 2021 году".format(ws.cell(ws.max_row-1, 1).value)
    # построение графиков
    # график "ДИАГРАММА ИЗМЕРЯЕМЫХ ПРАМЕТРОВ ПО НЕДЕЛЯМ"

    cats = Reference(ws, min_row=2, max_row=ws.max_row - 1, min_col=1, max_col=1)
    values = Reference(ws, min_row=1, max_row=ws.max_row - 1, min_col=2, max_col=8)
    # chart = LineChart()
    chart = BarChart()
    chart.y_axis.title = 'Параметры'
    chart.x_axis.title = 'Недели'
    chart.height = 10
    chart.width = 30
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "A{}".format(ws.max_row + 2))

    # график выработки ГПЭГ, СМ и общий
    ch1 = LineChart()
    cats = Reference(ws, min_row=2, max_row=ws.max_row - 1, min_col=1, max_col=1)
    values = Reference(ws, min_row=1, max_row=ws.max_row - 1, min_col=2, max_col=4)
    ch1.title = "ВЫРАБОТКА ЭЛЕКТРОЭНЕРГИИ"  # заголовок
    ch1.style = 13  # шрифт
    ch1.height = 10  # высота
    ch1.width = 20  # ширина
    ch1.x_axis.title = 'Недели'  # подпись оси х
    ch1.y_axis.title = 'кВт*ч'  # подпись оси у
    ch1.legend.position = 'r'  # позиция подписей данных справа
    ch1.add_data(values, titles_from_data=True)  # загрузка данных с заголовками столбцов
    ch1.set_categories(cats)  # загрузка подписи оси х
    ch1.series[0].graphicalProperties.line.solidFill = "85C1E9"  # цвет синий
    ch1.series[1].graphicalProperties.line.solidFill = "F7DC6F"  # цвет желтый
    ch1.series[2].graphicalProperties.line.solidFill = "EC7063"  # цвет красный
    ch1.series[0].graphicalProperties.solidFill = "85C1E9"  # цвет синий
    ch1.series[1].graphicalProperties.solidFill = "F7DC6F"  # цвет желтый
    ch1.series[2].graphicalProperties.solidFill = "EC7063"  # цвет красный
    ws.add_chart(ch1, "A{}".format(ws.max_row + 22))  # загрузка графика в ячейку

    # график ПОТРЕБЛЕНИЕ ЭЛЕКТРОЭНЕРГИИ
    ch2 = LineChart()
    cats = Reference(ws, min_row=2, max_row=ws.max_row - 1, min_col=1, max_col=1)
    values = Reference(ws, min_row=1, max_row=ws.max_row - 1, min_col=4, max_col=6)
    ch2.title = "ПОТРЕБЛЕНИЕ ЭЛЕКТРОЭНЕРГИИ"  # заголовок
    ch2.style = 13  # шрифт
    ch2.height = 10  # высота
    ch2.width = 20  # ширина
    ch2.x_axis.title = 'Недели'  # подпись оси х
    ch2.y_axis.title = 'кВт*ч'  # подпись оси у
    ch2.legend.position = 'r'  # позиция подписей данных справа
    ch2.add_data(values, titles_from_data=True)  # загрузка данных с заголовками столбцов
    ch2.set_categories(cats)  # загрузка подписи оси х
    ch2.series[0].graphicalProperties.line.solidFill = "EC7063"  # цвет линии красный
    ch2.series[1].graphicalProperties.line.solidFill = "F7DC6F"  # цвет линии желтый
    ch2.series[2].graphicalProperties.line.solidFill = "85C1E9"  # цвет линии синий
    ch2.series[0].graphicalProperties.solidFill = "EC7063"  # цвет заливки красный
    ch2.series[1].graphicalProperties.solidFill = "F7DC6F"  # цвет заливки желтый
    ch2.series[2].graphicalProperties.solidFill = "85C1E9"  # цвет заливки синий
    ws.add_chart(ch2, "A{}".format(ws.max_row + 42))  # загрузка графика в ячейку

    # график ПОТРЕБЛЕНИЕ ГАЗА
    ch3 = LineChart()
    cats = Reference(ws, min_row=2, max_row=ws.max_row - 1, min_col=1, max_col=1)
    values = Reference(ws, min_row=1, max_row=ws.max_row - 1, min_col=10, max_col=11)
    ch3.title = "ПОТРЕБЛЕНИЕ ГАЗА"  # заголовок
    ch3.style = 13  # шрифт
    ch3.height = 10  # высота
    ch3.width = 20  # ширина
    ch3.x_axis.title = 'Недели'  # подпись оси х
    ch3.y_axis.title = 'м3'  # подпись оси у
    ch3.legend.position = 'r'  # позиция подписей данных справа
    ch3.add_data(values, titles_from_data=True)  # загрузка данных с заголовками столбцов
    ch3.set_categories(cats)  # загрузка подписи оси х
    ch3.series[0].graphicalProperties.line.solidFill = "EC7063"  # цвет линии красный
    ch3.series[1].graphicalProperties.line.solidFill = "85C1E9"  # цвет линии синий
    ch3.series[0].graphicalProperties.solidFill = "EC7063"  # цвет заливки красный
    ch3.series[1].graphicalProperties.solidFill = "85C1E9"  # цвет заливки синий

    ch31 = LineChart()
    cats = Reference(ws, min_row=2, max_row=ws.max_row - 1, min_col=1, max_col=1)
    values1 = Reference(ws, min_row=1, max_row=ws.max_row - 1, min_col=8, max_col=8)
    ch31.title = "ПОТРЕБЛЕНИЕ ГАЗА"  # заголовок
    ch31.style = 13  # шрифт
    ch31.height = 10  # высота
    ch31.width = 20  # ширина
    ch31.x_axis.title = 'Недели'  # подпись оси х
    ch31.y_axis.title = 'м3'  # подпись оси у
    ch31.legend.position = 'r'  # позиция подписей данных справа
    ch31.add_data(values1, titles_from_data=True)  # загрузка данных с заголовками столбцов
    ch31.set_categories(cats)  # загрузка подписи оси х

    ch3 += ch31
    ws.add_chart(ch3, "A{}".format(ws.max_row + 62))  # загрузка графика в ячейку

    # график ЭФФЕКТИВНОСТЬ ВЫРАБОТКИ ЭЛЕКТРОЭНЕРГИИ БКЭУ
    ch4 = LineChart()
    cats = Reference(ws, min_row=2, max_row=ws.max_row - 1, min_col=1, max_col=1)
    values = Reference(ws, min_row=1, max_row=ws.max_row - 1, min_col=12, max_col=12)
    ch4.title = "ЭФФЕКТИВНОСТЬ ВЫРАБОТКИ ЭЛЕКТРОЭНЕРГИИ БКЭУ"  # заголовок
    ch4.style = 13  # шрифт
    ch4.height = 10  # высота
    ch4.width = 20  # ширина
    ch4.x_axis.title = 'Недели'  # подпись оси х
    ch4.y_axis.title = 'м3/кВт*ч'  # подпись оси у
    ch4.legend.position = 'r'  # позиция подписей данных справа
    ch4.add_data(values, titles_from_data=True)  # загрузка данных с заголовками столбцов
    ch4.set_categories(cats)  # загрузка подписи оси х
    ch4.series[0].graphicalProperties.line.solidFill = "28B463"  # цвет линии зеленый
    ch4.series[0].graphicalProperties.solidFill = "28B463"  # цвет заливки зеленый

    ws.add_chart(ch4, "A{}".format(ws.max_row + 82))  # загрузка графика в ячейку

    # график ВЫРАБОТКА ТЕПЛОВОЙ ЭНЕРГИИ
    ch5 = LineChart()
    cats = Reference(ws, min_row=2, max_row=ws.max_row - 1, min_col=1, max_col=1)
    values = Reference(ws, min_row=1, max_row=ws.max_row - 1, min_col=9, max_col=9)
    ch5.title = "ВЫРАБОТКА ТЕПЛОВОЙ ЭНЕРГИИ"  # заголовок
    ch5.style = 13  # шрифт
    ch5.height = 10  # высота
    ch5.width = 20  # ширина
    ch5.x_axis.title = 'Недели'  # подпись оси х
    ch5.y_axis.title = 'кВт*ч'  # подпись оси у
    ch5.legend.position = 'r'  # позиция подписей данных справа
    ch5.add_data(values, titles_from_data=True)  # загрузка данных с заголовками столбцов
    ch5.set_categories(cats)  # загрузка подписи оси х
    ch5.series[0].graphicalProperties.line.solidFill = "C0392B"  # цвет линии красный
    ch5.series[0].graphicalProperties.solidFill = "C0392B"  # цвет заливки красный

    ws.add_chart(ch5, "A{}".format(ws.max_row + 102))  # загрузка графика в ячейку

    # расчет суммы выработки ГПЭГ и СМ для построения общей диаграммы
    ws.cell(ws.max_row, 2).value = 0
    ws.cell(ws.max_row, 3).value = 0
    for i in range(2, ws.max_row + 1):
        if ws.cell(i, 1).value is None:
            ws.cell(i, 1).value = "Итого:"
        if ws.cell(i, 2).value is not None:
            ws.cell(ws.max_row, 2).value += ws.cell(i, 2).value
            ws.cell(ws.max_row, 2).number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[1]
        if ws.cell(i, 3).value is not None:
            ws.cell(ws.max_row, 3).value += ws.cell(i, 3).value
            ws.cell(ws.max_row, 3).number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[1]

    # построение общей диаграммы выработки ГПЭГ и СМ за отчетный период
    chart_itog = PieChart()
    labels = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=1)
    data = Reference(ws, min_col=2, max_col=3, min_row=ws.max_row, max_row=ws.max_row)
    chart_itog.title = "Выработка за отчетный период {} недель".format(ws.max_row - 2)
    chart_itog.style = 13  # шрифт
    chart_itog.height = 10  # высота
    chart_itog.width = 20  # ширина
    chart_itog.add_data(data, titles_from_data=True)  # загрузка данных с заголовками столбцов
    chart_itog.set_categories(labels)  # загрузка подписи оси х
    slice = DataPoint(idx=0, explosion=20)  # разделение пирога и сдвиг на 20 пунктов
    chart_itog.series[0].data_points = [slice]  # применение сдвига к первому значению
    ws.add_chart(chart_itog, "A{}".format(ws.max_row + 122))

    # график температур внутри блок-бокса
    ch6 = LineChart()
    cats = Reference(ws, min_row=2, max_row=ws.max_row - 1, min_col=1, max_col=1)
    values = Reference(ws, min_row=1, max_row=ws.max_row - 1, min_col=13, max_col=14)
    ch6.title = "ТЕМПЕРАТУРА ВНУТРИ БКЭУ"  # заголовок
    ch6.style = 13  # шрифт
    ch6.height = 10  # высота
    ch6.width = 20  # ширина
    ch6.x_axis.title = 'Недели'  # подпись оси х
    ch6.y_axis.title = '°C'  # подпись оси у
    ch6.legend.position = 'r'  # позиция подписей данных справа
    ch6.add_data(values, titles_from_data=True)  # загрузка данных с заголовками столбцов
    ch6.set_categories(cats)  # загрузка подписи оси х
    ch6.series[0].graphicalProperties.line.solidFill = "85C1E9"  # цвет синий
    ch6.series[1].graphicalProperties.line.solidFill = "EC7063"  # цвет красный
    ch6.series[0].graphicalProperties.solidFill = "85C1E9"  # цвет синий
    ch6.series[1].graphicalProperties.solidFill = "EC7063"  # цвет красный
    ws.add_chart(ch6, "A{}".format(ws.max_row + 142))  # загрузка графика в ячейку

    wb.save(path_save+'\\Графики.xlsx')  # сохранение таблицы в указанную директорию

    # ------ СОХРАНЕНИЕ ГРАФИКОВ ИЗ ТАБЛИЦЫ EXCEL В ФОРМАТЕ PNG-----------------
    # input_file = "C:/Razrab-10/python/ExcelWordIntegration/Test.xlsx"
    # output_image = "C:/Razrab-10/python/ExcelWordIntegration/"
    #
    # operation = win32com.client.Dispatch("Excel.Application")
    # operation.Visible = 0
    # operation.DisplayAlerts = 0
    #
    # workbook_2 = operation.Workbooks.Open(input_file)
    # sheet_2 = operation.Sheets(1)
    #
    # for x, chart in enumerate(sheet_2.Shapes):
    #     chart.Copy()
    #     image = ImageGrab.grabclipboard()
    #     image.save(output_image + "{}.png".format(x), 'png')
    #     pass
    #
    # workbook_2.Close(True)
    # operation.Quit()
    # ---------------------------------------------------------------------------- Сохранение графиков из таблицы в формате png


    # -------Генерация автоматического отчета в WORD------------------------
    template = DocxTemplate('temp6707.docx')

    #week = 17  # задаем номер недели для отчета
    # задаем начальные значения недельных параметров
    Wfull = 0
    Wcm = 0
    Wgpeg = 0
    Wcn = 0
    Wnagr = 0
    Qgvk = 0
    moto = 0
    Vgvk = 0
    Vgpeg = 0
    Vbkeu = 0
    Tmin = 0
    Tmax = 0
    # задаем начальные значения суммы параметров за период с начала ОПЭ
    Wfull_sum = 0
    Wcm_sum = 0
    Wgpeg_sum = 0
    Wcn_sum = 0
    Wnagr_sum = 0
    Qgvk_sum = 0
    Vgvk_sum = 0
    Vgpeg_sum = 0
    Vbkeu_sum = 0
    moto_sum = 0
    # задаем начальные значения дня и месяца начала и окончания недели
    d_start = 0
    m_start = 0
    d_end = 0
    m_end = 0
    # рассчитываем значения за неделю и сумму с начала ОПЭ
    for i in range(2, ws.max_row):
        Wfull_sum += ws.cell(i, 4).value
        Wcm_sum += ws.cell(i, 3).value
        Wgpeg_sum += ws.cell(i, 2).value
        Wcn_sum += ws.cell(i, 6).value
        Wnagr_sum += ws.cell(i, 5).value
        Qgvk_sum += ws.cell(i, 9).value
        Vgvk_sum += ws.cell(i, 11).value
        Vgpeg_sum += ws.cell(i, 10).value
        Vbkeu_sum += ws.cell(i, 8).value
        moto_sum += ws.cell(i, 7).value
        if ws.cell(i, 1).value == week:
            Wfull = ws.cell(i, 4).value
            Wcm = ws.cell(i, 3).value
            Wgpeg = ws.cell(i, 2).value
            Wcn = ws.cell(i, 6).value
            Wnagr = ws.cell(i, 5).value
            Qgvk = ws.cell(i, 9).value
            moto = ws.cell(i, 7).value
            Vgvk = ws.cell(i, 11).value
            Vgpeg = ws.cell(i, 10).value
            Vbkeu = ws.cell(i, 8).value
            Tmin = ws.cell(i, 13).value
            Tmax = ws.cell(i, 14).value
            d_start = ws.cell(i, 15).value
            m_start = ws.cell(i, 15).value
            d_end = ws.cell(i, 16).value
            m_end = ws.cell(i, 16).value
            break


    def month_name(num):  # функция возврата названия месяца по его номеру
        ru = ['января', 'февраля', 'марта', 'апреля', 'мая', 'июня', 'июля', 'августа', 'сентября', 'октября', 'ноября',
              'декабря']
        return ru[int(num) - 1]


    # Объявляем значения переменных, идентичных шаблону в документе word
    context = {
        'week': week,
        'year': d_end.strftime('%Y'),
        'Wfull': round(Wfull, 2),
        'Wcm': round(Wcm, 2),
        'Wgpeg': round(Wgpeg, 2),
        'Wcn': round(Wcn, 2),
        'Wnagr': round(Wnagr, 2),
        'Qgvk': round(Qgvk, 1),
        'moto': moto,
        'Vgvk': round(Vgvk, 2),
        'Vgpeg': round(Vgpeg, 2),
        'Vbkeu': round(Vbkeu, 2),
        'Tmin': Tmin,
        'Tmax': Tmax,
        'Wfull_sum': round(Wfull_sum, 1),
        'Wcm_sum': round(Wcm_sum, 1),
        'Wgpeg_sum': round(Wgpeg_sum, 1),
        'Wcn_sum': round(Wcn_sum, 1),
        'Wnagr_sum': round(Wnagr_sum, 1),
        'Qgvk_sum': round(Qgvk_sum),
        'Vgvk_sum': round(Vgvk_sum, 1),
        'Vgpeg_sum': round(Vgpeg_sum, 1),
        'Vbkeu_sum': round(Vbkeu_sum, 1),
        'moto_sum': moto_sum,
        'd_start': d_start.strftime('%d'),
        'm_start': month_name(m_start.strftime('%m')),
        'd_end': d_end.strftime('%d'),
        'm_end': month_name(m_end.strftime('%m'))
    }

    # создаем автоматизированный отчет
    template.render(context)
    template.save(path_save + '\\Еженедельный отчет по ОПЭ БКЭУ-{}.docx'.format(week))  # сохранение отчета с атоприсвоением номера недели
    file_path = path_save + '\\Еженедельный отчет по ОПЭ БКЭУ-{}.docx'.format(week)
    os.startfile(file_path)
    return "Отчет успешно сформирован"

# whatis()
